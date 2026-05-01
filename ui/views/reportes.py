from decimal import Decimal
from django.contrib.auth import get_user_model
from django.db.models import Q, Sum, Value, DecimalField, Count
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.views.generic import TemplateView
from ui.services.pdf import render_to_pdf

from finanzas.models import Comision
from finanzas.models import Pago
from polizas.models import Poliza
from cotizador.models import Cotizacion



User = get_user_model()

class ReporteMenuView(TemplateView):
    template_name = "ui/reportes/menu.html"

class ReporteComisionesView(TemplateView):
    template_name = "ui/reportes/comisiones.html"

    def get(self, request, *args, **kwargs):
        qs = self.get_queryset()

        if request.GET.get("export") == "excel":
            return self.export_excel(qs)
        if request.GET.get("export") == "pdf":
                return self.export_pdf(qs)

        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = (
            Comision.objects
            .select_related("poliza", "poliza__cliente", "agente")
            .order_by("-fecha_generacion", "-id")
        )

        q = self.request.GET.get("q", "").strip()
        agente_id = self.request.GET.get("agente", "").strip()
        estatus = self.request.GET.get("estatus", "").strip()
        desde = self.request.GET.get("desde", "").strip()
        hasta = self.request.GET.get("hasta", "").strip()

        if q:
            qs = qs.filter(
                Q(poliza__numero_poliza__icontains=q) |
                Q(poliza__cliente__nombre__icontains=q) |
                Q(agente__first_name__icontains=q) |
                Q(agente__last_name__icontains=q) |
                Q(agente__username__icontains=q)
            )

        if agente_id:
            qs = qs.filter(agente_id=agente_id)

        if estatus:
            qs = qs.filter(estatus=estatus)

        if desde:
            qs = qs.filter(fecha_generacion__gte=desde)

        if hasta:
            qs = qs.filter(fecha_generacion__lte=hasta)

        return qs

    def totalizar(self, qs, estatus=None):
        if estatus:
            qs = qs.filter(estatus=estatus)

        return qs.aggregate(
            total=Coalesce(
                Sum("monto_comision"),
                Value(Decimal("0.00"), output_field=DecimalField(max_digits=14, decimal_places=2)),
            )
        )["total"]

    def export_excel(self, qs):
        encabezados = [
            "Póliza",
            "Cliente",
            "Agente",
            "Porcentaje",
            "Base cálculo",
            "Monto comisión",
            "Estatus",
            "Fecha generación",
            "Fecha pago",
        ]

        filas = []

        for c in qs:
            filas.append([
                getattr(c.poliza, "numero_poliza", "") or c.poliza.id,
                str(c.poliza.cliente) if c.poliza and c.poliza.cliente else "",
                str(c.agente) if c.agente else "",
                float(c.porcentaje),
                float(c.base_calculo),
                float(c.monto_comision),
                c.get_estatus_display(),
                c.fecha_generacion.strftime("%d/%m/%Y") if c.fecha_generacion else "",
                c.fecha_pago.strftime("%d/%m/%Y") if c.fecha_pago else "",
            ])

        return generar_excel_response(
            nombre_archivo="reporte_comisiones",
            encabezados=encabezados,
            filas=filas,
            titulo="Reporte de Comisiones",
        )

    def export_pdf(self, qs):
        return render_to_pdf(
            "ui/reportes/pdf/comisiones_pdf.html",
            {
                "comisiones": qs,
                "total_pendiente": self.totalizar(qs, Comision.Estatus.PENDIENTE),
                "total_pagado": self.totalizar(qs, Comision.Estatus.PAGADA),
                "total_cancelado": self.totalizar(qs, Comision.Estatus.CANCELADA),
                "total_registros": qs.count(),
                # filtros
                "q": self.request.GET.get("q", "").strip(),
                "agente_id": self.request.GET.get("agente", "").strip(),
                "estatus": self.request.GET.get("estatus", "").strip(),
                "desde": self.request.GET.get("desde", "").strip(),
                "hasta": self.request.GET.get("hasta", "").strip(),
            },
            filename="reporte_comisiones.pdf",
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        qs = self.get_queryset()

        ctx["comisiones"] = qs

        ctx["agentes"] = (
            User.objects
            .filter(comisiones__isnull=False)
            .distinct()
            .order_by("first_name", "last_name", "username")
        )

        ctx["q"] = self.request.GET.get("q", "").strip()
        ctx["agente_id"] = self.request.GET.get("agente", "").strip()
        ctx["estatus"] = self.request.GET.get("estatus", "").strip()
        ctx["desde"] = self.request.GET.get("desde", "").strip()
        ctx["hasta"] = self.request.GET.get("hasta", "").strip()

        ctx["total_registros"] = qs.count()
        ctx["total_pendiente"] = self.totalizar(qs, Comision.Estatus.PENDIENTE)
        ctx["total_pagado"] = self.totalizar(qs, Comision.Estatus.PAGADA)
        ctx["total_cancelado"] = self.totalizar(qs, Comision.Estatus.CANCELADA)
        ctx["total_general"] = self.totalizar(qs)

        return ctx

User = get_user_model()

class ReporteCarteraVencidaView(TemplateView):
    template_name = "ui/reportes/cartera_vencida.html"

    def get(self, request, *args, **kwargs):
        qs = self.get_queryset()

        if request.GET.get("export") == "excel":
            return self.export_excel(qs)
        if request.GET.get("export") == "pdf":
                return self.export_pdf(qs)

        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = (
            Pago.objects
            .select_related(
                "poliza",
                "poliza__cliente",
                "poliza__agente",
                "poliza__aseguradora",
            )
            .filter(estatus=Pago.Estatus.VENCIDO)
            .order_by("fecha_programada", "id")
        )

        q = self.request.GET.get("q", "").strip()
        agente_id = self.request.GET.get("agente", "").strip()
        desde = self.request.GET.get("desde", "").strip()
        hasta = self.request.GET.get("hasta", "").strip()

        if q:
            qs = qs.filter(
                Q(poliza__numero_poliza__icontains=q) |
                Q(poliza__cliente__nombre__icontains=q) |
                Q(poliza__aseguradora__nombre__icontains=q) |
                Q(poliza__agente__first_name__icontains=q) |
                Q(poliza__agente__last_name__icontains=q) |
                Q(poliza__agente__username__icontains=q)
            )

        if agente_id:
            qs = qs.filter(poliza__agente_id=agente_id)

        if desde:
            qs = qs.filter(fecha_programada__gte=desde)

        if hasta:
            qs = qs.filter(fecha_programada__lte=hasta)

        return qs

    def totalizar(self, qs):
        return qs.aggregate(
            total=Coalesce(
                Sum("monto"),
                Value(Decimal("0.00"), output_field=DecimalField(max_digits=14, decimal_places=2)),
            )
        )["total"]

    def export_excel(self, qs):
        encabezados = [
            "Fecha vencimiento",
            "Póliza",
            "Cliente",
            "Aseguradora",
            "Agente",
            "Monto",
            "Estatus",
        ]

        filas = []

        for p in qs:
            filas.append([
                p.fecha_programada.strftime("%d/%m/%Y") if p.fecha_programada else "",
                getattr(p.poliza, "numero_poliza", "") or p.poliza.id,
                str(p.poliza.cliente) if p.poliza and p.poliza.cliente else "",
                str(p.poliza.aseguradora) if p.poliza and p.poliza.aseguradora else "",
                str(p.poliza.agente) if p.poliza and p.poliza.agente else "",
                float(p.monto),
                p.get_estatus_display(),
            ])

        return generar_excel_response(
            nombre_archivo="reporte_cartera_vencida",
            encabezados=encabezados,
            filas=filas,
            titulo="Reporte de Cartera Vencida",
        )

    def export_pdf(self, qs):
        return render_to_pdf(
            "ui/reportes/pdf/cartera_vencida_pdf.html",
            {
                "pagos": qs,
                "total_registros": qs.count(),
                "monto_total_vencido": self.totalizar(qs),
                # filtros
                "q": self.request.GET.get("q", "").strip(),
                "agente_id": self.request.GET.get("agente", "").strip(),
                "desde": self.request.GET.get("desde", "").strip(),
                "hasta": self.request.GET.get("hasta", "").strip(),

            },
            filename="reporte_cartera_vencida.pdf",
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        qs = self.get_queryset()

        ctx["pagos"] = qs

        agente_ids = (
            qs.exclude(poliza__agente__isnull=True)
            .values_list("poliza__agente_id", flat=True)
            .distinct()
        )

        ctx["agentes"] = (
            User.objects
            .filter(id__in=agente_ids)
            .order_by("first_name", "last_name", "username")
        )

        ctx["q"] = self.request.GET.get("q", "").strip()
        ctx["agente_id"] = self.request.GET.get("agente", "").strip()
        ctx["desde"] = self.request.GET.get("desde", "").strip()
        ctx["hasta"] = self.request.GET.get("hasta", "").strip()

        ctx["total_registros"] = qs.count()
        ctx["monto_total_vencido"] = self.totalizar(qs)

        ctx["clientes_vencidos"] = (
            qs.values("poliza__cliente_id")
            .distinct()
            .count()
        )

        ctx["polizas_vencidas"] = (
            qs.values("poliza_id")
            .distinct()
            .count()
        )

        return ctx

User = get_user_model()

class ReporteRenovacionesView(TemplateView):
    template_name = "ui/reportes/renovaciones.html"

    def get(self, request, *args, **kwargs):
        qs = self.get_queryset()

        if request.GET.get("export") == "excel":
            return self.export_excel(qs)
        if request.GET.get("export") == "pdf":
                return self.export_pdf(qs)

        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        today = timezone.localdate()
        default_hasta = today + timezone.timedelta(days=30)

        qs = (
            Poliza.objects
            .select_related("cliente", "aseguradora", "agente")
            .filter(
                estatus=Poliza.Estatus.VIGENTE,
                vigencia_hasta__gte=today,
            )
            .order_by("vigencia_hasta", "id")
        )

        q = self.request.GET.get("q", "").strip()
        agente_id = self.request.GET.get("agente", "").strip()
        aseguradora_id = self.request.GET.get("aseguradora", "").strip()
        desde = self.request.GET.get("desde", "").strip()
        hasta = self.request.GET.get("hasta", "").strip()

        if not desde:
            qs = qs.filter(vigencia_hasta__gte=today)
        else:
            qs = qs.filter(vigencia_hasta__gte=desde)

        if not hasta:
            qs = qs.filter(vigencia_hasta__lte=default_hasta)
        else:
            qs = qs.filter(vigencia_hasta__lte=hasta)

        if q:
            qs = qs.filter(
                Q(numero_poliza__icontains=q) |
                Q(cliente__nombre__icontains=q) |
                Q(aseguradora__nombre__icontains=q) |
                Q(agente__first_name__icontains=q) |
                Q(agente__last_name__icontains=q) |
                Q(agente__username__icontains=q)
            )

        if agente_id:
            qs = qs.filter(agente_id=agente_id)

        if aseguradora_id:
            qs = qs.filter(aseguradora_id=aseguradora_id)

        return qs

    def totalizar_prima(self, qs):
        return qs.aggregate(
            total=Coalesce(
                Sum("prima_total"),
                Value(Decimal("0.00"), output_field=DecimalField(max_digits=14, decimal_places=2)),
            )
        )["total"]

    def export_excel(self, qs):
        encabezados = [
            "Fecha vencimiento",
            "Póliza",
            "Cliente",
            "Aseguradora",
            "Agente",
            "Prima total",
            "Estatus",
        ]

        filas = []

        for p in qs:
            filas.append([
                p.vigencia_hasta.strftime("%d/%m/%Y") if p.vigencia_hasta else "",
                getattr(p, "numero_poliza", "") or p.id,
                str(p.cliente) if p.cliente else "",
                str(p.aseguradora) if p.aseguradora else "",
                str(p.agente) if p.agente else "",
                float(p.prima_total) if p.prima_total else 0,
                p.get_estatus_display(),
            ])

        return generar_excel_response(
            nombre_archivo="reporte_renovaciones",
            encabezados=encabezados,
            filas=filas,
            titulo="Reporte de Renovaciones Próximas",
        )

    def export_pdf(self, qs):
        return render_to_pdf(
            "ui/reportes/pdf/renovaciones_pdf.html",
            {
                "polizas": qs,
                "total_registros": qs.count(),
                "prima_total_renovar": self.totalizar_prima(qs),
                # filtros
                "q": self.request.GET.get("q", "").strip(),
                "agente_id": self.request.GET.get("agente", "").strip(),
                "aseguradora": self.request.GET.get("aseguradora", "").strip(),
                "desde": self.request.GET.get("desde", "").strip(),
                "hasta": self.request.GET.get("hasta", "").strip(),

            },
            filename="reporte_renovaciones.pdf",
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        today = timezone.localdate()
        default_hasta = today + timezone.timedelta(days=30)
        next_7 = today + timezone.timedelta(days=7)

        qs = self.get_queryset()

        ctx["polizas"] = qs

        agente_ids = (
            qs.exclude(agente__isnull=True)
            .values_list("agente_id", flat=True)
            .distinct()
        )

        ctx["agentes"] = (
            User.objects
            .filter(id__in=agente_ids)
            .order_by("first_name", "last_name", "username")
        )

        ctx["aseguradoras"] = (
            Poliza.objects
            .filter(
                estatus=Poliza.Estatus.VIGENTE,
                vigencia_hasta__gte=today,
                vigencia_hasta__lte=default_hasta,
            )
            .exclude(aseguradora__isnull=True)
            .values("aseguradora_id", "aseguradora__nombre")
            .distinct()
            .order_by("aseguradora__nombre")
        )

        ctx["q"] = self.request.GET.get("q", "").strip()
        ctx["agente_id"] = self.request.GET.get("agente", "").strip()
        ctx["aseguradora_id"] = self.request.GET.get("aseguradora", "").strip()
        ctx["desde"] = self.request.GET.get("desde", "").strip() or today.isoformat()
        ctx["hasta"] = self.request.GET.get("hasta", "").strip() or default_hasta.isoformat()

        ctx["total_registros"] = qs.count()
        ctx["prima_total_renovar"] = self.totalizar_prima(qs)

        ctx["proximas_7d"] = qs.filter(
            vigencia_hasta__gte=today,
            vigencia_hasta__lte=next_7,
        ).count()

        ctx["proximas_30d"] = qs.filter(
            vigencia_hasta__gte=today,
            vigencia_hasta__lte=default_hasta,
        ).count()

        ctx["today"] = today
        ctx["next_7"] = next_7
        ctx["default_hasta"] = default_hasta

        return ctx


User = get_user_model()

class ReporteProduccionAgenteView(TemplateView):
    template_name = "ui/reportes/produccion_agente.html"

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") == "excel":
            return self.export_excel()
        if request.GET.get("export") == "pdf":
            return self.export_pdf()

        return super().get(request, *args, **kwargs)

    def get_queryset(self):
        qs = (
            Poliza.objects
            .select_related("agente", "cliente", "aseguradora")
            .exclude(agente__isnull=True)
            .order_by("agente__first_name", "agente__last_name", "agente__username")
        )

        agente_id = self.request.GET.get("agente", "").strip()
        desde = self.request.GET.get("desde", "").strip()
        hasta = self.request.GET.get("hasta", "").strip()
        estatus = self.request.GET.get("estatus", "").strip()

        if agente_id:
            qs = qs.filter(agente_id=agente_id)

        if desde:
            qs = qs.filter(created_at__date__gte=desde)

        if hasta:
            qs = qs.filter(created_at__date__lte=hasta)

        if estatus:
            qs = qs.filter(estatus=estatus)

        return qs

    def export_excel(self):
        ctx = self.get_context_data()
        filas_ctx = ctx["filas"]

        encabezados = [
            "Agente",
            "Pólizas",
            "Prima neta",
            "Prima total",
            "Comisiones generadas",
            "Comisiones pagadas",
            "Comisiones pendientes",
        ]

        filas = []

        for r in filas_ctx:
            filas.append([
                r["agente_nombre"],
                r["polizas_count"],
                float(r["prima_neta_total"]),
                float(r["prima_total_total"]),
                float(r["com_generadas"]),
                float(r["com_pagadas"]),
                float(r["com_pendientes"]),
            ])

        # Totales (footer en Excel)
        filas.append([
            "TOTAL",
            ctx["total_polizas"],
            float(ctx["total_prima_neta"]),
            float(ctx["total_prima_total"]),
            float(ctx["total_comisiones"]),
            "",
            "",
        ])

        return generar_excel_response(
            nombre_archivo="reporte_produccion_agente",
            encabezados=encabezados,
            filas=filas,
            titulo="Reporte de Producción por Agente",
        )

    def export_pdf(self):
        ctx = self.get_context_data()
        # agregar filtros al mismo contexto
        ctx.update({
            "agente_id": self.request.GET.get("agente", "").strip(),
            "estatus": self.request.GET.get("estatus", "").strip(),
            "desde": self.request.GET.get("desde", "").strip(),
            "hasta": self.request.GET.get("hasta", "").strip(),
        })

        return render_to_pdf(
            "ui/reportes/pdf/produccion_agente_pdf.html",
            ctx,
            filename="reporte_produccion_agente.pdf",
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        polizas_qs = self.get_queryset()

        agente_id = self.request.GET.get("agente", "").strip()
        desde = self.request.GET.get("desde", "").strip()
        hasta = self.request.GET.get("hasta", "").strip()
        estatus = self.request.GET.get("estatus", "").strip()

        agentes = (
            User.objects
            .filter(polizas__isnull=False)
            .distinct()
            .order_by("first_name", "last_name", "username")
        )

        produccion = (
            polizas_qs
            .values(
                "agente_id",
                "agente__first_name",
                "agente__last_name",
                "agente__username",
            )
            .annotate(
                polizas_count=Count("id", distinct=True),
                prima_neta_total=Coalesce(
                    Sum("prima_neta"),
                    Value(Decimal("0.00"), output_field=DecimalField(max_digits=14, decimal_places=2)),
                ),
                prima_total_total=Coalesce(
                    Sum("prima_total"),
                    Value(Decimal("0.00"), output_field=DecimalField(max_digits=14, decimal_places=2)),
                ),
            )
            .order_by("-prima_total_total", "agente__first_name")
        )

        filas = []

        for row in produccion:
            agente = row["agente_id"]

            comisiones_qs = Comision.objects.filter(agente_id=agente)

            if desde:
                comisiones_qs = comisiones_qs.filter(fecha_generacion__gte=desde)

            if hasta:
                comisiones_qs = comisiones_qs.filter(fecha_generacion__lte=hasta)

            com_generadas = comisiones_qs.aggregate(
                total=Coalesce(
                    Sum("monto_comision"),
                    Value(Decimal("0.00"), output_field=DecimalField(max_digits=14, decimal_places=2)),
                )
            )["total"]

            com_pagadas = comisiones_qs.filter(
                estatus=Comision.Estatus.PAGADA
            ).aggregate(
                total=Coalesce(
                    Sum("monto_comision"),
                    Value(Decimal("0.00"), output_field=DecimalField(max_digits=14, decimal_places=2)),
                )
            )["total"]

            com_pendientes = comisiones_qs.filter(
                estatus=Comision.Estatus.PENDIENTE
            ).aggregate(
                total=Coalesce(
                    Sum("monto_comision"),
                    Value(Decimal("0.00"), output_field=DecimalField(max_digits=14, decimal_places=2)),
                )
            )["total"]

            nombre = f"{row['agente__first_name']} {row['agente__last_name']}".strip()
            if not nombre:
                nombre = row["agente__username"]

            filas.append({
                "agente_id": agente,
                "agente_nombre": nombre,
                "polizas_count": row["polizas_count"],
                "prima_neta_total": row["prima_neta_total"],
                "prima_total_total": row["prima_total_total"],
                "com_generadas": com_generadas,
                "com_pagadas": com_pagadas,
                "com_pendientes": com_pendientes,
            })

        total_polizas = sum(f["polizas_count"] for f in filas)
        total_prima_neta = sum(f["prima_neta_total"] for f in filas) if filas else Decimal("0.00")
        total_prima_total = sum(f["prima_total_total"] for f in filas) if filas else Decimal("0.00")
        total_comisiones = sum(f["com_generadas"] for f in filas) if filas else Decimal("0.00")

        ctx["filas"] = filas
        ctx["agentes"] = agentes

        ctx["agente_id"] = agente_id
        ctx["desde"] = desde
        ctx["hasta"] = hasta
        ctx["estatus"] = estatus

        ctx["total_agentes"] = len(filas)
        ctx["total_polizas"] = total_polizas
        ctx["total_prima_neta"] = total_prima_neta
        ctx["total_prima_total"] = total_prima_total
        ctx["total_comisiones"] = total_comisiones

        return ctx

User = get_user_model()

class ReporteConversionAgenteView(TemplateView):
    template_name = "ui/reportes/conversion_agente.html"

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") == "excel":
            return self.export_excel()

        if request.GET.get("export") == "pdf":
            return self.export_pdf()

        return super().get(request, *args, **kwargs)

    def export_pdf(self):
        ctx = self.get_context_data()
        # agregar filtros al mismo contexto
        ctx.update({
            "agente_id": self.request.GET.get("agente", "").strip(),
            "desde": self.request.GET.get("desde", "").strip(),
            "hasta": self.request.GET.get("hasta", "").strip(),
        })

        return render_to_pdf(
            "ui/reportes/pdf/conversion_agente_pdf.html",
            ctx,
            filename="reporte_conversion_agente.pdf",
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        agente_id = self.request.GET.get("agente", "").strip()
        desde = self.request.GET.get("desde", "").strip()
        hasta = self.request.GET.get("hasta", "").strip()

        cot_qs = Cotizacion.objects.all()
        pol_qs = Poliza.objects.all()

        if agente_id:
            cot_qs = cot_qs.filter(owner_id=agente_id)
            pol_qs = pol_qs.filter(agente_id=agente_id)

        if desde:
            cot_qs = cot_qs.filter(created_at__date__gte=desde)
            pol_qs = pol_qs.filter(created_at__date__gte=desde)

        if hasta:
            cot_qs = cot_qs.filter(created_at__date__lte=hasta)
            pol_qs = pol_qs.filter(created_at__date__lte=hasta)

        cot_agg = (
            cot_qs
            .values("owner_id", "owner__first_name", "owner__last_name", "owner__username")
            .annotate(cotizaciones=Count("id"))
        )

        pol_agg = (
            pol_qs
            .values("agente_id")
            .annotate(polizas=Count("id"))
        )

        pol_map = {p["agente_id"]: p["polizas"] for p in pol_agg}

        filas = []

        for row in cot_agg:
            agente_id = row["owner_id"]

            cotizaciones = row["cotizaciones"]
            polizas = pol_map.get(agente_id, 0)

            conversion = (polizas / cotizaciones * 100) if cotizaciones else 0

            nombre = f"{row['owner__first_name']} {row['owner__last_name']}".strip()
            if not nombre:
                nombre = row["owner__username"]

            filas.append({
                "agente_id": agente_id,
                "agente_nombre": nombre,
                "cotizaciones": cotizaciones,
                "polizas": polizas,
                "conversion": round(conversion, 2),
            })

        total_cot = sum(f["cotizaciones"] for f in filas)
        total_pol = sum(f["polizas"] for f in filas)
        conversion_global = (total_pol / total_cot * 100) if total_cot else 0

        agentes = (
            User.objects
            .filter(id__in=[f["agente_id"] for f in filas])
            .order_by("first_name", "last_name", "username")
        )

        ctx["filas"] = sorted(filas, key=lambda x: x["conversion"], reverse=True)
        ctx["agentes"] = agentes

        ctx["agente_id"] = self.request.GET.get("agente", "")
        ctx["desde"] = desde
        ctx["hasta"] = hasta

        ctx["total_cot"] = total_cot
        ctx["total_pol"] = total_pol
        ctx["conversion_global"] = round(conversion_global, 2)

        return ctx

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Exportar a Excel

def generar_excel_response(nombre_archivo, encabezados, filas, titulo=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    row_num = 1

    if titulo:
        ws.cell(row=row_num, column=1, value=titulo)
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
        row_num += 2

    header_fill = PatternFill("solid", fgColor="E9ECEF")

    for col_num, encabezado in enumerate(encabezados, 1):
        cell = ws.cell(row=row_num, column=col_num, value=encabezado)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_num += 1

    for fila in filas:
        for col_num, valor in enumerate(fila, 1):
            ws.cell(row=row_num, column=col_num, value=valor)
        row_num += 1

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}.xlsx"'

    wb.save(response)
    return response
